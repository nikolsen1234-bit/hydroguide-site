"""Parse Solar_calculator.xlsx into HydroConfigData + results."""

from io import BytesIO

import openpyxl

from app.models.schemas import (
    BatteryParams,
    CommunicationParams,
    DieselGeneratorParams,
    EnergyBalanceResult,
    FacilityParams,
    FuelCellParams,
    HydroConfigData,
    MonthlyEnergyBalance,
    MonthlyIrradiation,
    OperationsParams,
    OtherSettings,
    PowerBudgetItem,
    RecommendedConfig,
    SolarParams,
    TcoComparison,
)


def _yes_no_to_bool(value) -> bool | None:
    """Convert 'Yes'/'No'/None to bool."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("yes", "ja", "true", "1"):
        return True
    if s in ("no", "nei", "false", "0"):
        return False
    return None


def _safe_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _cell(ws, row: int, col: int):
    """Get cell value by 1-indexed row/col."""
    return ws.cell(row=row, column=col).value


def parse_input_sheet(ws) -> tuple[HydroConfigData, list[str]]:
    """Parse the Input sheet into HydroConfigData.

    Returns (config, validation_notes).
    """
    notes: list[str] = []

    # Section 1: Communication (rows 5–8, answers in column B)
    communication = CommunicationParams(
        has_4g_coverage=_yes_no_to_bool(_cell(ws, 5, 2)),
        has_nbiot_coverage=_yes_no_to_bool(_cell(ws, 6, 2)),
        has_line_of_sight=_yes_no_to_bool(_cell(ws, 7, 2)),
        requires_two_way_control=_yes_no_to_bool(_cell(ws, 8, 2)),
    )

    # Section 2: Facility (rows 10–17)
    facility = FacilityParams(
        release_method=str(_cell(ws, 10, 2) or ""),
        has_fish_passage=_yes_no_to_bool(_cell(ws, 11, 2)),
        minimum_flow_ls=_safe_float(_cell(ws, 12, 2)),
        low_conductivity=_yes_no_to_bool(_cell(ws, 13, 2)),
        ice_problems=_yes_no_to_bool(_cell(ws, 14, 2)),
        difficult_access=_yes_no_to_bool(_cell(ws, 15, 2)),
        linear_flow=_yes_no_to_bool(_cell(ws, 16, 2)),
        sediment_or_surge=_yes_no_to_bool(_cell(ws, 17, 2)),
    )

    # Section 3: Operations (rows 19–21)
    operations = OperationsParams(
        inspections_per_year=_safe_int(_cell(ws, 19, 2)),
        battery_bank_ah=_safe_float(_cell(ws, 20, 2)),
        zero_emission_desired=_yes_no_to_bool(_cell(ws, 21, 2)),
    )

    # Solar (rows 31–34, values in column C)
    solar = SolarParams(
        panel_wattage_wp=_safe_float(_cell(ws, 31, 3)),
        panel_count=_safe_int(_cell(ws, 32, 3)) or 1,
        system_efficiency=_safe_float(_cell(ws, 33, 3)) or 0.8,
        lifespan_years=_safe_int(_cell(ws, 34, 3)) or 25,
    )

    # Battery (rows 38–40, values in column C)
    battery = BatteryParams(
        voltage_v=_safe_float(_cell(ws, 38, 3)) or 12.8,
        max_dod=_safe_float(_cell(ws, 39, 3)) or 0.8,
        cycle_lifespan=_safe_int(_cell(ws, 40, 3)) or 6000,
    )

    # Fuel cell (rows 44–49, values in column C)
    fuel_cell = FuelCellParams(
        purchase_cost_kr=_safe_float(_cell(ws, 44, 3)),
        power_w=_safe_float(_cell(ws, 45, 3)),
        fuel_consumption_l_kwh=_safe_float(_cell(ws, 46, 3)),
        fuel_price_kr_l=_safe_float(_cell(ws, 47, 3)),
        lifespan_hours=_safe_int(_cell(ws, 48, 3)),
        annual_maintenance_kr=_safe_float(_cell(ws, 49, 3)),
    )

    # Diesel generator (rows 52–57, values in column C)
    diesel = DieselGeneratorParams(
        purchase_cost_kr=_safe_float(_cell(ws, 52, 3)),
        power_w=_safe_float(_cell(ws, 53, 3)),
        fuel_consumption_l_kwh=_safe_float(_cell(ws, 54, 3)),
        fuel_price_kr_l=_safe_float(_cell(ws, 55, 3)),
        lifespan_hours=_safe_int(_cell(ws, 56, 3)),
        annual_maintenance_kr=_safe_float(_cell(ws, 57, 3)),
    )

    # Other settings (rows 60–62, values in column C)
    other = OtherSettings(
        co2_factor_methanol=_safe_float(_cell(ws, 60, 3)) or 1.088,
        co2_factor_diesel=_safe_float(_cell(ws, 61, 3)) or 2.68,
        assessment_horizon_years=_safe_int(_cell(ws, 62, 3)) or 10,
    )

    # Monthly irradiation (rows 66–77, values in column C)
    irr = MonthlyIrradiation(
        jan=_safe_float(_cell(ws, 66, 3)) or 0,
        feb=_safe_float(_cell(ws, 67, 3)) or 0,
        mar=_safe_float(_cell(ws, 68, 3)) or 0,
        apr=_safe_float(_cell(ws, 69, 3)) or 0,
        may=_safe_float(_cell(ws, 70, 3)) or 0,
        jun=_safe_float(_cell(ws, 71, 3)) or 0,
        jul=_safe_float(_cell(ws, 72, 3)) or 0,
        aug=_safe_float(_cell(ws, 73, 3)) or 0,
        sep=_safe_float(_cell(ws, 74, 3)) or 0,
        oct=_safe_float(_cell(ws, 75, 3)) or 0,
        nov=_safe_float(_cell(ws, 76, 3)) or 0,
        dec=_safe_float(_cell(ws, 77, 3)) or 0,
    )

    config = HydroConfigData(
        communication=communication,
        facility=facility,
        operations=operations,
        solar=solar,
        battery=battery,
        fuel_cell=fuel_cell,
        diesel_generator=diesel,
        other_settings=other,
        monthly_irradiation=irr,
    )

    return config, notes


def parse_power_budget_sheet(ws) -> list[PowerBudgetItem]:
    """Parse the Power Budget sheet into a list of equipment items.

    Reads rows 3+ until an empty name or 'Total' row.
    Columns: A=ON/OFF, B=Unit name, C=Power(W), D=Current(A),
             E=Wh/day, F=Ah/day, G=Wh/week, H=Ah/week
    """
    items: list[PowerBudgetItem] = []
    for row in range(3, ws.max_row + 1):
        name = _cell(ws, row, 2)
        if not name or str(name).strip().lower() == "total":
            # Skip empty rows but keep going (might have gaps)
            if str(_cell(ws, row, 1) or "").strip().lower() == "total":
                break
            if not name:
                continue

        enabled = _cell(ws, row, 1)
        if isinstance(enabled, bool):
            enabled = enabled
        else:
            enabled = True

        items.append(PowerBudgetItem(
            enabled=enabled,
            name=str(name).strip(),
            power_w=_safe_float(_cell(ws, row, 3)) or 0,
            consumption_wh_day=_safe_float(_cell(ws, row, 5)) or 0,
            consumption_ah_day=_safe_float(_cell(ws, row, 6)) or 0,
            consumption_wh_week=_safe_float(_cell(ws, row, 7)) or 0,
            consumption_ah_week=_safe_float(_cell(ws, row, 8)) or 0,
        ))

    return items


def parse_results_sheet(
    ws,
) -> tuple[RecommendedConfig, EnergyBalanceResult | None, TcoComparison | None]:
    """Parse the Results sheet for recommended config, energy balance, and TCO."""

    # Recommended configuration (rows 5–15, values in column B)
    recommended = RecommendedConfig(
        measurement_method=str(_cell(ws, 5, 2) or ""),
        communication=str(_cell(ws, 6, 2) or ""),
        logger_recommendation=str(_cell(ws, 7, 2) or ""),
        energy_monitoring=str(_cell(ws, 8, 2) or ""),
        secondary_source=str(_cell(ws, 9, 2) or ""),
        secondary_source_power_w=_safe_float(_cell(ws, 10, 2)),
        solar_panel_count=_safe_int(_cell(ws, 11, 2)),
        battery_ah=_safe_float(_cell(ws, 12, 2)),
        fuel_consumption_l_yr=_safe_float(_cell(ws, 13, 2)),
        ice_adaptation=str(_cell(ws, 14, 2) or ""),
        operational_requirements=str(_cell(ws, 15, 2) or ""),
    )

    # Energy balance (rows 27–38 = Jan–Dec, row 39 = SUM)
    month_names = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    ]
    monthly: list[MonthlyEnergyBalance] = []
    for i, name in enumerate(month_names):
        row = 27 + i
        monthly.append(MonthlyEnergyBalance(
            month=name,
            days=_safe_int(_cell(ws, row, 2)) or 30,
            solar_production_kwh=_safe_float(_cell(ws, row, 3)) or 0,
            energy_balance_kwh=_safe_float(_cell(ws, row, 4)) or 0,
            generator_hours=_safe_float(_cell(ws, row, 5)) or 0,
            fuel_liters=_safe_float(_cell(ws, row, 6)) or 0,
            fuel_cost_kr=_safe_float(_cell(ws, row, 7)) or 0,
        ))

    # SUM row (39)
    energy_balance = EnergyBalanceResult(
        monthly=monthly,
        total_solar_production_kwh=_safe_float(_cell(ws, 39, 3)) or 0,
        total_energy_balance_kwh=_safe_float(_cell(ws, 39, 4)) or 0,
        total_generator_hours=_safe_float(_cell(ws, 39, 5)) or 0,
        total_fuel_liters=_safe_float(_cell(ws, 39, 6)) or 0,
        total_fuel_cost_kr=_safe_float(_cell(ws, 39, 7)) or 0,
    )

    # TCO comparison (rows 43–47)
    tco = TcoComparison(
        fuel_cell_purchase_kr=_safe_float(_cell(ws, 43, 3)) or 0,
        fuel_cell_operating_kr_yr=_safe_float(_cell(ws, 44, 3)) or 0,
        fuel_cell_maintenance_kr_yr=_safe_float(_cell(ws, 45, 3)) or 0,
        fuel_cell_tco_kr=_safe_float(_cell(ws, 47, 3)) or 0,
        diesel_purchase_kr=_safe_float(_cell(ws, 43, 4)) or 0,
        diesel_operating_kr_yr=_safe_float(_cell(ws, 44, 4)) or 0,
        diesel_maintenance_kr_yr=_safe_float(_cell(ws, 45, 4)) or 0,
        diesel_tco_kr=_safe_float(_cell(ws, 47, 4)) or 0,
        assessment_horizon_years=_safe_int(_cell(ws, 46, 3)) or 10,
        recommended_source=(
            "fuel_cell"
            if (_safe_float(_cell(ws, 47, 3)) or float("inf"))
            < (_safe_float(_cell(ws, 47, 4)) or float("inf"))
            else "diesel"
        ),
    )

    return recommended, energy_balance, tco


def parse_excel(file_bytes: bytes) -> tuple[
    HydroConfigData,
    RecommendedConfig | None,
    EnergyBalanceResult | None,
    TcoComparison | None,
    list[str],
]:
    """Parse a Solar_calculator.xlsx file.

    Args:
        file_bytes: Raw bytes of the .xlsx file.

    Returns:
        (config, recommended_config, energy_balance, tco, validation_notes)
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    notes: list[str] = []

    # Parse Input sheet
    if "Input" not in wb.sheetnames:
        raise ValueError("Excel file must contain an 'Input' sheet")
    config, input_notes = parse_input_sheet(wb["Input"])
    notes.extend(input_notes)

    # Parse Power Budget sheet
    if "Power Budget" in wb.sheetnames:
        config.power_budget = parse_power_budget_sheet(wb["Power Budget"])
    else:
        notes.append("No 'Power Budget' sheet found — power budget will be empty")

    # Parse Results sheet
    recommended = None
    energy_balance = None
    tco = None
    if "Results" in wb.sheetnames:
        recommended, energy_balance, tco = parse_results_sheet(wb["Results"])
    else:
        notes.append("No 'Results' sheet found — no reference results to compare")

    return config, recommended, energy_balance, tco, notes
